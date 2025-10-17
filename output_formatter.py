"""
Output formatting module for gene misannotation detection results.

This module handles generation of multiple output formats:
- CSV: Reorganized columns with summary statistics
- TSV: Tab-separated for easy terminal/text viewing
- Excel: Multi-sheet workbook with formatting and documentation

All outputs include documentation of the 10 AA overlap threshold used for filtering.
"""

import pandas as pd
import os
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Color
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import ColorScaleRule
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel output will be skipped.")
    print("Install with: pip install openpyxl")

import scoring


# ============================================================================
# COLUMN ORGANIZATION
# ============================================================================

# Define the priority order for columns in output files
COLUMN_ORDER = [
    # Scores and identifiers (most important)
    "composite_score",
    "fused_protein",
    "fused_product",

    # New analysis columns
    "organism_count",
    "theorized_intron_length_bp",

    # Gene part information
    "gene_1",
    "product_1",
    "gene_1_len_aa",
    "gene_2",
    "product_2",
    "gene_2_len_aa",
    "fused_gene_len_aa",

    # Subject match information
    "subject_id",
    "subject_title",
    "subject_length",

    # Alignment quality metrics
    "query_coverage",
    "subject_coverage",
    "percentage_of_identical_matches",
    "bit_score",
    "expected_value",

    # Alignment position ranges (normalized 0-1 scale) - PRIORITIZED
    "alignment_range_in_query_norm",
    "alignment_range_in_subject_norm",

    # Alignment position ranges (absolute amino acid positions)
    "alignment_range_in_query",
    "alignment_range_in_subject",
    "alignment_length",

    # Additional statistics
    "number_of_identical_matches",
    "number_of_mismatches",
    "query_length",
    "query_title",
]


def reorganize_columns(df):
    """
    Reorganize dataframe columns according to priority order.

    Args:
        df (pd.DataFrame): Input dataframe

    Returns:
        pd.DataFrame: Dataframe with reorganized columns
    """
    # Get columns that exist in both the desired order and the dataframe
    existing_ordered_cols = [col for col in COLUMN_ORDER if col in df.columns]

    # Get any remaining columns not in our predefined order
    remaining_cols = [col for col in df.columns if col not in COLUMN_ORDER]

    # Combine: ordered columns first, then any extras
    final_column_order = existing_ordered_cols + remaining_cols

    return df[final_column_order]


def generate_summary_statistics(df_fused, df_control, organism_name):
    """
    Generate summary statistics for the results.

    Args:
        df_fused (pd.DataFrame): Fused hits dataframe with scores
        df_control (pd.DataFrame): Control hits dataframe
        organism_name (str): Name of organism being analyzed

    Returns:
        dict: Summary statistics
    """
    stats = {
        "organism": organism_name,
        "analysis_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_fused_hits": len(df_fused),
        "total_control_hits": len(df_control),
        "overlap_threshold_aa": 10,  # Hardcoded threshold documented here
    }

    if len(df_fused) > 0:
        stats["avg_score"] = df_fused["composite_score"].mean()
        stats["median_score"] = df_fused["composite_score"].median()
        stats["max_score"] = df_fused["composite_score"].max()
        stats["min_score"] = df_fused["composite_score"].min()

        # Score distribution
        stats["high_confidence_hits"] = len(df_fused[df_fused["composite_score"] >= 70])
        stats["medium_confidence_hits"] = len(df_fused[(df_fused["composite_score"] >= 40) &
                                                       (df_fused["composite_score"] < 70)])
        stats["low_confidence_hits"] = len(df_fused[df_fused["composite_score"] < 40])

    return stats


def create_summary_header(stats, scoring_info):
    """
    Create header text with summary statistics and methodology.

    Args:
        stats (dict): Summary statistics
        scoring_info (dict): Scoring configuration info

    Returns:
        str: Formatted header text
    """
    header = f"""# Gene Misannotation Detection Results
# Organism: {stats['organism']}
# Analysis Date: {stats['analysis_date']}
#
# SUMMARY STATISTICS:
# Total Fused Hits: {stats['total_fused_hits']}
# Total Control Hits: {stats['total_control_hits']}
# Overlap Threshold: {stats['overlap_threshold_aa']} amino acids on each side of fusion junction
#
"""

    if stats['total_fused_hits'] > 0:
        header += f"""# SCORE DISTRIBUTION:
# High Confidence (score ≥ 70): {stats['high_confidence_hits']}
# Medium Confidence (40 ≤ score < 70): {stats['medium_confidence_hits']}
# Low Confidence (score < 40): {stats['low_confidence_hits']}
#
# Score Range: {stats['min_score']:.2f} - {stats['max_score']:.2f}
# Mean Score: {stats['avg_score']:.2f}
# Median Score: {stats['median_score']:.2f}
#
"""

    header += f"""# SCORING METHODOLOGY (0-100 scale):
# Query Coverage Weight: {scoring_info['weights']['query_coverage']} - {scoring_info['descriptions']['query_coverage']}
# Bit Score Improvement Weight: {scoring_info['weights']['bitscore_improvement']} - {scoring_info['descriptions']['bitscore_improvement']}
# Overlap Equilibrium Weight: {scoring_info['weights']['overlap_equilibrium']} - {scoring_info['descriptions']['overlap_equilibrium']}
# Organism Count Weight: {scoring_info['weights']['organism_count']} - {scoring_info['descriptions']['organism_count']}
#
"""

    return header


def write_csv_with_header(df, filepath, header_text):
    """
    Write CSV with custom header comments.

    Args:
        df (pd.DataFrame): Dataframe to write
        filepath (str): Output file path
        header_text (str): Header comment text
    """
    with open(filepath, 'w') as f:
        f.write(header_text)
        df.to_csv(f, index=False)


def write_tsv_with_header(df, filepath, header_text):
    """
    Write TSV with custom header comments.

    Args:
        df (pd.DataFrame): Dataframe to write
        filepath (str): Output file path
        header_text (str): Header comment text
    """
    with open(filepath, 'w') as f:
        f.write(header_text)
        df.to_csv(f, index=False, sep='\t')


def auto_adjust_column_widths(worksheet, df):
    """
    Automatically adjust column widths based on content.

    Args:
        worksheet: openpyxl worksheet object
        df: pandas DataFrame with the data
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    # Calculate length, accounting for header formatting
                    cell_length = len(str(cell.value))
                    if max_length < cell_length:
                        max_length = cell_length
            except:
                pass

        # Set width with some padding (multiply by 1.2 for extra space)
        # Minimum width of 10, maximum of 50 to keep things readable
        adjusted_width = min(max(max_length * 1.2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def create_excel_workbook(df_fused, df_control, stats, scoring_info, output_path):
    """
    Create Excel workbook with multiple sheets and formatting.

    Args:
        df_fused (pd.DataFrame): Fused hits dataframe
        df_control (pd.DataFrame): Control hits dataframe
        stats (dict): Summary statistics
        scoring_info (dict): Scoring configuration
        output_path (str): Output file path
    """
    if not EXCEL_AVAILABLE:
        print("Skipping Excel output (openpyxl not installed)")
        return

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # ========================================================================
    # SHEET 1: Top Hits (score ≥ 70)
    # ========================================================================
    ws_top = wb.create_sheet("Top Hits")

    df_top = df_fused[df_fused["composite_score"] >= 70].copy() if len(df_fused) > 0 else pd.DataFrame()

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df_top, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_top.cell(row=r_idx, column=c_idx, value=value)

            # Header formatting
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")

    # Apply color gradient to score column (green = high, red = low)
    if len(df_top) > 0:
        score_col = 'A'  # composite_score is first column
        # ColorScale: red (0) -> yellow (50) -> green (100)
        color_scale = ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',   # Red
            mid_type='num', mid_value=50, mid_color='FFEB84',        # Yellow
            end_type='num', end_value=100, end_color='63BE7B'        # Green
        )
        ws_top.conditional_formatting.add(f'{score_col}2:{score_col}{len(df_top) + 1}', color_scale)

    # Auto-adjust column widths
    auto_adjust_column_widths(ws_top, df_top)

    # ========================================================================
    # SHEET 2: All Fused Hits
    # ========================================================================
    ws_all = wb.create_sheet("All Fused Hits")

    for r_idx, row in enumerate(dataframe_to_rows(df_fused, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")

    # Apply color gradient to score column (green = high, red = low)
    if len(df_fused) > 0:
        score_col = 'A'  # composite_score is first column
        # ColorScale: red (0) -> yellow (50) -> green (100)
        color_scale = ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',   # Red
            mid_type='num', mid_value=50, mid_color='FFEB84',        # Yellow
            end_type='num', end_value=100, end_color='63BE7B'        # Green
        )
        ws_all.conditional_formatting.add(f'{score_col}2:{score_col}{len(df_fused) + 1}', color_scale)

    # Auto-adjust column widths
    auto_adjust_column_widths(ws_all, df_fused)

    # ========================================================================
    # SHEET 3: Control Hits
    # ========================================================================
    ws_control = wb.create_sheet("Control Hits")

    for r_idx, row in enumerate(dataframe_to_rows(df_control, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_control.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    auto_adjust_column_widths(ws_control, df_control)

    # ========================================================================
    # SHEET 4: Scoring Methodology
    # ========================================================================
    ws_info = wb.create_sheet("Scoring Methodology")

    # Write summary statistics
    ws_info.cell(1, 1, "GENE MISANNOTATION DETECTION RESULTS").font = Font(bold=True, size=14)
    ws_info.cell(3, 1, "ANALYSIS INFORMATION").font = Font(bold=True)
    ws_info.cell(4, 1, "Organism:")
    ws_info.cell(4, 2, stats['organism'])
    ws_info.cell(5, 1, "Analysis Date:")
    ws_info.cell(5, 2, stats['analysis_date'])
    ws_info.cell(6, 1, "Total Fused Hits:")
    ws_info.cell(6, 2, stats['total_fused_hits'])
    ws_info.cell(7, 1, "Total Control Hits:")
    ws_info.cell(7, 2, stats['total_control_hits'])
    ws_info.cell(8, 1, "Overlap Threshold (AA):")
    ws_info.cell(8, 2, stats['overlap_threshold_aa'])

    if stats['total_fused_hits'] > 0:
        ws_info.cell(10, 1, "SCORE DISTRIBUTION").font = Font(bold=True)
        ws_info.cell(11, 1, "High Confidence (≥70):")
        ws_info.cell(11, 2, stats['high_confidence_hits'])
        ws_info.cell(12, 1, "Medium Confidence (40-69):")
        ws_info.cell(12, 2, stats['medium_confidence_hits'])
        ws_info.cell(13, 1, "Low Confidence (<40):")
        ws_info.cell(13, 2, stats['low_confidence_hits'])
        ws_info.cell(15, 1, "Mean Score:")
        ws_info.cell(15, 2, f"{stats['avg_score']:.2f}")
        ws_info.cell(16, 1, "Median Score:")
        ws_info.cell(16, 2, f"{stats['median_score']:.2f}")

    ws_info.cell(18, 1, "SCORING METHODOLOGY").font = Font(bold=True)
    ws_info.cell(19, 1, "Score Range: 0-100 (higher = more likely true misannotation)")

    row = 21
    ws_info.cell(row, 1, "Component").font = Font(bold=True)
    ws_info.cell(row, 2, "Weight").font = Font(bold=True)
    ws_info.cell(row, 3, "Description").font = Font(bold=True)

    row += 1
    ws_info.cell(row, 1, "Query Coverage")
    ws_info.cell(row, 2, scoring_info['weights']['query_coverage'])
    ws_info.cell(row, 3, scoring_info['descriptions']['query_coverage'])

    row += 1
    ws_info.cell(row, 1, "Bit Score Improvement")
    ws_info.cell(row, 2, scoring_info['weights']['bitscore_improvement'])
    ws_info.cell(row, 3, scoring_info['descriptions']['bitscore_improvement'])

    row += 1
    ws_info.cell(row, 1, "Overlap Equilibrium")
    ws_info.cell(row, 2, scoring_info['weights']['overlap_equilibrium'])
    ws_info.cell(row, 3, scoring_info['descriptions']['overlap_equilibrium'])

    row += 1
    ws_info.cell(row, 1, "Organism Count")
    ws_info.cell(row, 2, scoring_info['weights']['organism_count'])
    ws_info.cell(row, 3, scoring_info['descriptions']['organism_count'])

    # Adjust column widths
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 15
    ws_info.column_dimensions['C'].width = 60

    # Save workbook
    wb.save(output_path)


def generate_all_outputs(df_fused, df_control, organism_name, output_folder):
    """
    Generate all output formats (CSV, TSV, Excel).

    Args:
        df_fused (pd.DataFrame): Fused hits dataframe with scores
        df_control (pd.DataFrame): Control hits dataframe
        organism_name (str): Name of organism being analyzed
        output_folder (str): Output directory path
    """
    # Reorganize columns for better readability
    df_fused_organized = reorganize_columns(df_fused)
    df_control_organized = reorganize_columns(df_control)

    # Generate summary statistics
    stats = generate_summary_statistics(df_fused_organized, df_control_organized, organism_name)
    scoring_info = scoring.get_scoring_info()

    # Create header text
    header_text = create_summary_header(stats, scoring_info)

    # ========================================================================
    # Generate CSV outputs
    # ========================================================================
    print("Generating CSV outputs...")

    # All fused hits (sorted by score)
    write_csv_with_header(
        df_fused_organized,
        os.path.join(output_folder, "fused_hits.csv"),
        header_text
    )

    # High confidence hits
    df_high = df_fused_organized[df_fused_organized["composite_score"] >= 70]
    if len(df_high) > 0:
        write_csv_with_header(
            df_high,
            os.path.join(output_folder, "high_confidence_hits.csv"),
            header_text.replace("Total Fused Hits", f"High Confidence Hits (score ≥ 70)")
        )

    # Control hits
    write_csv_with_header(
        df_control_organized,
        os.path.join(output_folder, "control_hits.csv"),
        "# Control Hits - Individual gene part alignments\n"
    )

    # ========================================================================
    # Generate TSV outputs
    # ========================================================================
    print("Generating TSV outputs...")

    write_tsv_with_header(
        df_fused_organized,
        os.path.join(output_folder, "fused_hits.tsv"),
        header_text
    )

    write_tsv_with_header(
        df_control_organized,
        os.path.join(output_folder, "control_hits.tsv"),
        "# Control Hits - Individual gene part alignments\n"
    )

    # ========================================================================
    # Generate Excel workbook
    # ========================================================================
    print("Generating Excel workbook...")

    create_excel_workbook(
        df_fused_organized,
        df_control_organized,
        stats,
        scoring_info,
        os.path.join(output_folder, "misannotation_results.xlsx")
    )

    print(f"Output generation complete. Files saved to: {output_folder}")
    print(f"  - CSV: fused_hits.csv, high_confidence_hits.csv, control_hits.csv")
    print(f"  - TSV: fused_hits.tsv, control_hits.tsv")
    if EXCEL_AVAILABLE:
        print(f"  - Excel: misannotation_results.xlsx")
    print(f"\nSummary: {stats['total_fused_hits']} fused hits, " +
          f"{stats['high_confidence_hits'] if 'high_confidence_hits' in stats else 0} high confidence")
